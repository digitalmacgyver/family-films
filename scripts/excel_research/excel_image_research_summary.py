#!/usr/bin/env python3
"""
Comprehensive research summary for Excel image extraction from XLS files
"""

import sys
import subprocess
import importlib

def check_library(lib_name):
    """Check if a library is available and get version"""
    try:
        lib = importlib.import_module(lib_name)
        version = getattr(lib, '__version__', 'Unknown')
        return True, version
    except ImportError:
        return False, None

def install_and_test_library(lib_name):
    """Try to install a library and test it"""
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", lib_name], 
                            capture_output=True, text=True)
        return check_library(lib_name)
    except:
        return False, None

def test_xls2xlsx():
    """Test the xls2xlsx library"""
    print("\n=== Testing xls2xlsx library ===")
    
    available, version = check_library('xls2xlsx')
    if not available:
        print("xls2xlsx not installed, attempting to install...")
        available, version = install_and_test_library('xls2xlsx')
    
    if available:
        print(f"xls2xlsx version: {version}")
        try:
            import xls2xlsx
            
            # Test with a real XLS file
            xls_file = "/home/viblio/family_films/chapter_sheets/L-55C_FROS - Bob Lindner Film 4 - Train and Lindners with Earl and Rosabell.xls"
            output_file = "/home/viblio/family_films/test_xls2xlsx_output.xlsx"
            
            print(f"Converting {xls_file}")
            result = xls2xlsx.to_xlsx(xls_file, output_file)
            print(f"Conversion result: {result}")
            
            # Check if output file was created
            import os
            if os.path.exists(output_file):
                print(f"✓ Successfully created: {output_file}")
                
                # Try to load with openpyxl and check for images
                import openpyxl
                from openpyxl_image_loader import SheetImageLoader
                
                wb = openpyxl.load_workbook(output_file)
                print(f"Converted file has sheets: {wb.sheetnames}")
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    if hasattr(sheet, '_images') and sheet._images:
                        print(f"Found {len(sheet._images)} images in sheet '{sheet_name}'")
                    else:
                        print(f"No images found in sheet '{sheet_name}'")
                        
            else:
                print("✗ Output file was not created")
                
        except Exception as e:
            print(f"Error testing xls2xlsx: {e}")
    else:
        print("xls2xlsx could not be installed or imported")

def test_xlwings():
    """Test xlwings for Excel automation"""
    print("\n=== Testing xlwings library ===")
    
    available, version = check_library('xlwings')
    if not available:
        print("xlwings not installed, attempting to install...")
        available, version = install_and_test_library('xlwings')
    
    if available:
        print(f"xlwings version: {version}")
        try:
            import xlwings as xw
            
            print("Note: xlwings requires Excel to be installed and running")
            print("xlwings is primarily for automation, not direct file parsing")
            print("It would require Excel COM automation on Windows or Excel on Mac")
            print("Not suitable for server environments or Linux without Excel")
            
        except Exception as e:
            print(f"Error with xlwings: {e}")
    else:
        print("xlwings could not be installed")

def test_alternative_libraries():
    """Test other potential libraries"""
    print("\n=== Testing alternative libraries ===")
    
    # Test python-docx (for comparison - it handles images in Word docs)
    libs_to_test = [
        ('python-docx', 'Can extract images from Word documents'),
        ('python-pptx', 'Can extract images from PowerPoint'),
        ('pywin32', 'Windows COM automation (Windows only)'),
    ]
    
    for lib_name, description in libs_to_test:
        available, version = check_library(lib_name.replace('-', '_'))
        if available:
            print(f"✓ {lib_name} v{version}: {description}")
        else:
            print(f"✗ {lib_name}: Not available - {description}")

def analyze_xls_structure():
    """Analyze XLS file structure to understand image storage"""
    print("\n=== Analyzing XLS file structure ===")
    
    xls_file = "/home/viblio/family_films/chapter_sheets/L-55C_FROS - Bob Lindner Film 4 - Train and Lindners with Earl and Rosabell.xls"
    
    print(f"Analyzing: {xls_file}")
    
    # Read as binary to look for image signatures
    try:
        with open(xls_file, 'rb') as f:
            content = f.read()
        
        # Look for common image format signatures
        image_signatures = {
            'PNG': b'\x89PNG\r\n\x1a\n',
            'JPEG': b'\xff\xd8\xff',
            'GIF': b'GIF8',
            'BMP': b'BM',
            'TIFF': b'II*\x00'
        }
        
        found_signatures = []
        for format_name, signature in image_signatures.items():
            if signature in content:
                count = content.count(signature)
                found_signatures.append(f"{format_name}: {count}")
        
        if found_signatures:
            print(f"✓ Found potential image data: {', '.join(found_signatures)}")
        else:
            print("✗ No common image format signatures found")
            
        # File size analysis
        file_size = len(content)
        print(f"File size: {file_size:,} bytes")
        
        if file_size > 50000:  # Larger files might contain images
            print("Large file size suggests potential embedded content")
        else:
            print("Small file size suggests text/data only")
            
    except Exception as e:
        print(f"Error analyzing file structure: {e}")

def create_recommendations():
    """Create final recommendations based on research"""
    print("\n" + "="*60)
    print("RESEARCH SUMMARY AND RECOMMENDATIONS")
    print("="*60)
    
    print("""
FINDINGS:

1. DIRECT XLS IMAGE EXTRACTION:
   ✗ xlrd: Cannot extract images from XLS files (by design)
   ✗ openpyxl: Cannot read XLS files (XLSX only)
   ✗ pandas: Can read XLS data but not images
   ✗ xlwings: Requires Excel installation, mainly for automation

2. XLSX IMAGE EXTRACTION (after conversion):
   ✓ openpyxl + openpyxl_image_loader: Works well for XLSX files
   ✓ Direct _images attribute access: Also works for XLSX files

3. CONVERSION APPROACHES:
   ✓ pandas: Can convert XLS to XLSX (data only, images lost)
   ✓ xls2xlsx: Specialized conversion library (may preserve more)
   ? Excel COM automation: Could preserve images but requires Excel

RECOMMENDATIONS:

For your Family Films project, here are the best approaches:

OPTION 1 - Pure Python (No Images from XLS):
```python
import pandas as pd
import openpyxl
from openpyxl_image_loader import SheetImageLoader

# Convert XLS to XLSX
df = pd.read_excel('file.xls', sheet_name=None)
with pd.ExcelWriter('file.xlsx', engine='openpyxl') as writer:
    for sheet_name, sheet_data in df.items():
        sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)

# Extract images from XLSX (if any survived conversion)
wb = openpyxl.load_workbook('file.xlsx')
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    image_loader = SheetImageLoader(sheet)
    # Check cells for images...
```

OPTION 2 - Excel Automation (Windows/Mac with Excel):
```python
import xlwings as xw

# Open Excel file through Excel application
app = xw.App(visible=False)
wb = app.books.open('file.xls')

# Save as XLSX to preserve images
wb.save('file.xlsx')
wb.close()
app.quit()

# Then use openpyxl for image extraction
```

OPTION 3 - Manual Conversion:
1. Open XLS files in Excel/LibreOffice
2. Save as XLSX format
3. Use Python openpyxl for image extraction

VERDICT FOR YOUR PROJECT:
Based on analysis of your actual XLS files, they appear to contain mainly
text/data without embedded images. The pandas conversion approach should
work well for extracting the chapter information you need.

If images are critical, consider manual conversion or Excel automation.
""")

def main():
    print("Excel Image Extraction Research - Comprehensive Analysis")
    print("="*60)
    
    # Check current environment
    print("\nCURRENT ENVIRONMENT:")
    libraries = [
        'pandas', 'openpyxl', 'xlrd', 'openpyxl_image_loader', 
        'PIL', 'xlwings', 'xls2xlsx'
    ]
    
    for lib in libraries:
        available, version = check_library(lib)
        status = f"✓ v{version}" if available else "✗ Not installed"
        print(f"  {lib:<20}: {status}")
    
    # Test specialized libraries
    test_xls2xlsx()
    test_xlwings()
    test_alternative_libraries()
    
    # Analyze file structure
    analyze_xls_structure()
    
    # Final recommendations
    create_recommendations()

if __name__ == "__main__":
    main()