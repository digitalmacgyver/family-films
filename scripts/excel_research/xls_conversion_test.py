#!/usr/bin/env python3
"""
Focused test for XLS to XLSX conversion and image extraction
"""

import pandas as pd
import openpyxl
from openpyxl_image_loader import SheetImageLoader
import xlrd
import os
import sys
from pathlib import Path

def test_xls_conversion_and_images(xls_file):
    """Test converting XLS to XLSX and then extracting any images"""
    print(f"\n=== Testing XLS conversion and image extraction ===")
    print(f"Source file: {xls_file}")
    
    if not os.path.exists(xls_file):
        print(f"File not found: {xls_file}")
        return False
    
    # Step 1: Test xlrd to understand XLS content
    print("\n1. Analyzing XLS with xlrd:")
    try:
        workbook = xlrd.open_workbook(xls_file)
        print(f"   Successfully opened with xlrd")
        print(f"   Sheets: {workbook.sheet_names()}")
        
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            print(f"   Sheet '{sheet_name}': {sheet.nrows} rows x {sheet.ncols} cols")
            
            # Sample some data
            if sheet.nrows > 0 and sheet.ncols > 0:
                sample_data = []
                for row in range(min(3, sheet.nrows)):
                    row_data = []
                    for col in range(min(3, sheet.ncols)):
                        try:
                            cell_value = sheet.cell_value(row, col)
                            row_data.append(str(cell_value)[:20])  # Truncate long values
                        except:
                            row_data.append("ERROR")
                    sample_data.append(row_data)
                print(f"   Sample data: {sample_data}")
                
    except Exception as e:
        print(f"   Error with xlrd: {e}")
        return False
    
    # Step 2: Convert to XLSX
    print("\n2. Converting XLS to XLSX:")
    xlsx_file = xls_file.replace('.xls', '_converted.xlsx')
    
    try:
        # Try different pandas approaches
        print("   Attempting basic pandas conversion...")
        df_dict = pd.read_excel(xls_file, sheet_name=None, header=None)
        
        print(f"   Read {len(df_dict)} sheets successfully")
        for sheet_name, df in df_dict.items():
            print(f"   Sheet '{sheet_name}': {df.shape[0]} rows x {df.shape[1]} cols")
        
        # Write to XLSX with proper engine
        print("   Writing to XLSX format...")
        with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
            for sheet_name, df in df_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        print(f"   Successfully created: {xlsx_file}")
        
    except Exception as e:
        print(f"   Error during conversion: {e}")
        
        # Try alternative approach with xlrd engine
        print("   Trying alternative approach with xlrd engine...")
        try:
            df_dict = pd.read_excel(xls_file, sheet_name=None, engine='xlrd', header=None)
            with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
                for sheet_name, df in df_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            print(f"   Successfully created with xlrd engine: {xlsx_file}")
        except Exception as e2:
            print(f"   Alternative approach also failed: {e2}")
            return False
    
    # Step 3: Check for images in converted XLSX
    print("\n3. Checking for images in converted XLSX:")
    
    if not os.path.exists(xlsx_file):
        print("   Converted file not found, skipping image check")
        return False
    
    try:
        workbook = openpyxl.load_workbook(xlsx_file)
        print(f"   Loaded converted XLSX with sheets: {workbook.sheetnames}")
        
        images_found = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\n   Checking sheet '{sheet_name}' for images:")
            
            # Method 1: openpyxl_image_loader
            try:
                image_loader = SheetImageLoader(sheet)
                
                # Check common cells where images might be
                test_cells = []
                # Generate test cells (A1-J10)
                for col in 'ABCDEFGHIJ':
                    for row in range(1, 11):
                        test_cells.append(f'{col}{row}')
                
                for cell in test_cells:
                    try:
                        if image_loader.image_in(cell):
                            image = image_loader.get(cell)
                            if image:
                                output_path = f"/home/viblio/family_films/extracted_{sheet_name}_{cell}.png"
                                image.save(output_path)
                                print(f"     ✓ Found and saved image from {cell} to {output_path}")
                                images_found += 1
                    except:
                        continue  # Silent fail for empty cells
                        
            except Exception as e:
                print(f"     Error with image_loader: {e}")
            
            # Method 2: Direct _images attribute check
            if hasattr(sheet, '_images') and sheet._images:
                print(f"     Found {len(sheet._images)} images via _images attribute")
                for i, img in enumerate(sheet._images):
                    try:
                        # Different approaches to get image data
                        if hasattr(img, '_data'):
                            img_data = img._data()
                        elif hasattr(img, 'ref'):
                            print(f"     Image {i} reference: {img.ref}")
                            continue
                        else:
                            print(f"     Image {i} attributes: {[attr for attr in dir(img) if not attr.startswith('_')]}")
                            continue
                            
                        output_path = f"/home/viblio/family_films/extracted_{sheet_name}_direct_{i}.png"
                        with open(output_path, 'wb') as f:
                            f.write(img_data)
                        print(f"     ✓ Saved image {i} to {output_path}")
                        images_found += 1
                        
                    except Exception as e:
                        print(f"     Error saving image {i}: {e}")
            else:
                print("     No images found via _images attribute")
        
        if images_found == 0:
            print("\n   ⚠ No images found in converted file")
            print("   This means either:")
            print("     - The original XLS file had no images")
            print("     - Images were lost during XLS->XLSX conversion")
            print("     - Images are embedded in a format not detectable by these methods")
        else:
            print(f"\n   ✅ Successfully found and extracted {images_found} images")
            
    except Exception as e:
        print(f"   Error checking for images: {e}")
        return False
    
    return True

def main():
    """Test with actual XLS files from the project"""
    xls_files = [
        "/home/viblio/family_films/chapter_sheets/L-55C_FROS - Bob Lindner Film 4 - Train and Lindners with Earl and Rosabell.xls",
        "/home/viblio/family_films/chapter_sheets/57-PT_FROS - Josephine Southwest Trip Grand Canyon Bryce Canyon Zion Calico Ghost Town.xls",
        "/home/viblio/family_films/chapter_sheets/L-55A_FROS - Bob Lindner Film 2 - House Fire Boat Trip Locks.xls"
    ]
    
    print("XLS to XLSX Conversion and Image Extraction Test")
    print("=" * 60)
    
    for xls_file in xls_files:
        if os.path.exists(xls_file):
            print(f"\n{'='*60}")
            print(f"Testing file: {os.path.basename(xls_file)}")
            test_xls_conversion_and_images(xls_file)
            break  # Test just the first available file for now
        else:
            print(f"File not found: {xls_file}")

if __name__ == "__main__":
    main()