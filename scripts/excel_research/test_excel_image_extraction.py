#!/usr/bin/env python3
"""
Test script for Excel image extraction research
This script tests various approaches to extract images from Excel files
"""

import pandas as pd
import openpyxl
from openpyxl_image_loader import SheetImageLoader
import xlrd
import os
import sys
from pathlib import Path

def test_openpyxl_xlsx_images(xlsx_file):
    """Test image extraction from XLSX files using openpyxl and openpyxl_image_loader"""
    print(f"\n=== Testing openpyxl image extraction from {xlsx_file} ===")
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(xlsx_file)
        print(f"Successfully loaded workbook with sheets: {workbook.sheetnames}")
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\nAnalyzing sheet: {sheet_name}")
            
            # Check for images using openpyxl_image_loader
            try:
                image_loader = SheetImageLoader(sheet)
                print("Image loader created successfully")
                
                # Try to find images in various cells
                test_cells = ['A1', 'B1', 'C1', 'A2', 'B2', 'C2', 'A3', 'B3', 'C3']
                images_found = 0
                
                for cell in test_cells:
                    try:
                        if image_loader.image_in(cell):
                            image = image_loader.get(cell)
                            if image:
                                output_path = f"/home/viblio/family_films/extracted_image_{sheet_name}_{cell}.png"
                                image.save(output_path)
                                print(f"  - Found and saved image from {cell} to {output_path}")
                                images_found += 1
                    except Exception as e:
                        # Silent fail for cells without images
                        pass
                
                if images_found == 0:
                    print("  - No images found in tested cells")
                    
            except Exception as e:
                print(f"  - Error with image loader: {e}")
                
            # Check for embedded objects directly
            if hasattr(sheet, '_images') and sheet._images:
                print(f"  - Found {len(sheet._images)} embedded images via _images attribute")
                for i, img in enumerate(sheet._images):
                    try:
                        img_data = img._data()
                        output_path = f"/home/viblio/family_films/extracted_image_{sheet_name}_direct_{i}.png"
                        with open(output_path, 'wb') as f:
                            f.write(img_data)
                        print(f"    - Saved image {i} to {output_path}")
                    except Exception as e:
                        print(f"    - Error saving image {i}: {e}")
            else:
                print("  - No images found via _images attribute")
                
    except Exception as e:
        print(f"Error loading XLSX file: {e}")
        return False
    
    return True

def test_openpyxl_xls_compatibility(xls_file):
    """Test if openpyxl can handle XLS files (it shouldn't)"""
    print(f"\n=== Testing openpyxl XLS compatibility with {xls_file} ===")
    
    try:
        workbook = openpyxl.load_workbook(xls_file)
        print("Unexpected: openpyxl opened XLS file!")
        return True
    except Exception as e:
        print(f"Expected error: {e}")
        return False

def test_xlrd_image_support(xls_file):
    """Test xlrd capabilities with XLS files (no image support expected)"""
    print(f"\n=== Testing xlrd with {xls_file} ===")
    
    try:
        workbook = xlrd.open_workbook(xls_file)
        print(f"Successfully opened XLS with xlrd")
        print(f"Sheet names: {workbook.sheet_names()}")
        
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            print(f"\nSheet '{sheet_name}': {sheet.nrows} rows x {sheet.ncols} cols")
            
            # xlrd doesn't support images, but let's check what attributes are available
            attrs = [attr for attr in dir(sheet) if not attr.startswith('_')]
            print(f"Available sheet attributes: {attrs[:10]}...")  # Show first 10
            
            # Check if there are any image-related methods (there won't be)
            image_attrs = [attr for attr in attrs if 'image' in attr.lower() or 'picture' in attr.lower()]
            if image_attrs:
                print(f"Image-related attributes found: {image_attrs}")
            else:
                print("No image-related attributes found (expected)")
                
    except Exception as e:
        print(f"Error opening XLS with xlrd: {e}")
        return False
    
    return True

def convert_xls_to_xlsx(xls_file, xlsx_file):
    """Convert XLS to XLSX using pandas"""
    print(f"\n=== Converting {xls_file} to {xlsx_file} using pandas ===")
    
    try:
        # Read XLS file
        print("Reading XLS file with pandas...")
        df = pd.read_excel(xls_file, sheet_name=None, header=None)  # Read all sheets
        
        print(f"Found {len(df)} sheets: {list(df.keys())}")
        
        # Write to XLSX
        print("Writing to XLSX format...")
        with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
            for sheet_name, sheet_data in df.items():
                sheet_data.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        print(f"Successfully converted to {xlsx_file}")
        return True
        
    except Exception as e:
        print(f"Error converting XLS to XLSX: {e}")
        return False

def create_test_xlsx_with_image():
    """Create a test XLSX file with an embedded image for testing"""
    print("\n=== Creating test XLSX file with image ===")
    
    try:
        # Create a simple test image (1x1 red pixel PNG)
        import io
        from PIL import Image
        
        # Create a small test image
        img = Image.new('RGB', (100, 100), color='red')
        test_img_path = "/home/viblio/family_films/test_image.png"
        img.save(test_img_path)
        
        # Create workbook and add image
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # Add some data
        ws['A1'] = "Test data"
        ws['B1'] = "With image below"
        
        # Add image
        from openpyxl.drawing.image import Image as OpenpyxlImage
        img = OpenpyxlImage(test_img_path)
        ws.add_image(img, 'A3')
        
        test_xlsx_path = "/home/viblio/family_films/test_file_with_image.xlsx"
        wb.save(test_xlsx_path)
        
        print(f"Created test XLSX file: {test_xlsx_path}")
        return test_xlsx_path
        
    except Exception as e:
        print(f"Error creating test XLSX: {e}")
        return None

def main():
    print("Excel Image Extraction Research")
    print("=" * 50)
    
    # Check if PIL is available for test image creation
    try:
        from PIL import Image
        pil_available = True
    except ImportError:
        print("PIL not available - will skip test image creation")
        pil_available = False
    
    # Create a test XLSX file with image if PIL is available
    test_xlsx_path = None
    if pil_available:
        test_xlsx_path = create_test_xlsx_with_image()
    
    # Test with created XLSX file
    if test_xlsx_path and os.path.exists(test_xlsx_path):
        test_openpyxl_xlsx_images(test_xlsx_path)
    
    # Look for existing Excel files in the directory
    current_dir = Path("/home/viblio/family_films")
    xls_files = list(current_dir.glob("*.xls"))
    xlsx_files = list(current_dir.glob("*.xlsx"))
    
    print(f"\nFound {len(xls_files)} XLS files and {len(xlsx_files)} XLSX files in directory")
    
    # Test with first XLS file if available
    if xls_files:
        xls_file = str(xls_files[0])
        print(f"Testing with: {xls_file}")
        
        # Test xlrd
        test_xlrd_image_support(xls_file)
        
        # Test openpyxl compatibility
        test_openpyxl_xls_compatibility(xls_file)
        
        # Test conversion
        converted_file = xls_file.replace('.xls', '_converted.xlsx')
        if convert_xls_to_xlsx(xls_file, converted_file):
            # Test image extraction from converted file
            test_openpyxl_xlsx_images(converted_file)
    
    # Test with existing XLSX files
    for xlsx_file in xlsx_files[:2]:  # Test first 2 XLSX files
        test_openpyxl_xlsx_images(str(xlsx_file))

if __name__ == "__main__":
    main()